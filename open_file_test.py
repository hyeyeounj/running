import os
import json
import pprint
from openpyxl import Workbook
from openpyxl import load_workbook
import sys  

reload(sys)  
sys.setdefaultencoding('utf8')

pwd = os.getcwd()

wb_total = Workbook()
wb_total_s = wb_total.active

cell_idx = 1

for path, dirs, files in os.walk(pwd):
	for file in files:
		if os.path.splitext(file)[1].lower() == '.xlsx':
		# 	if file.endswith(".xlsx"):
			filename = os.path.join(path, file)
			print 'FOLDER : ' + os.path.splitext(file)[0]
			wb = load_workbook(filename, data_only=True)
			ws = wb.worksheets[0]

			print ws['A3'].value or "ZZZ"
			print ws['B3'].value or "ZZZ" 

			wb_total_s['A1'] = ws['H4'].value
			# wb_total_s.append([1, 2, 3])

			wb_total.save("xxxxxx.xlsx")