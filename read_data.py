import os
import json
import pprint
from openpyxl import Workbook
from openpyxl import load_workbook
import sys  

reload(sys)  
sys.setdefaultencoding('utf8')

pwd = os.getcwd()

wb_total = Workbook()
wb_total_s = wb_total.active

cell_idx = 1
count = 1

def char_range(c1, c2):
	for c in xrange(ord(c1), ord(c2)+1):
		yield chr(c)

def scale_a_pre():
	idx = 3
	while idx < 26 :
		for col in char_range('B', 'Z'):
			wb_total_s[col +str(cell_idx)] = ws['B' + str(idx)].value or "XXX"
			print "A PRE: " + col + str(cell_idx) + " = " + "B" + str(idx)
			idx += 1
			
		
	for col_a in char_range('A', 'B'):
		for col_b in char_range('A', 'Z'):
			if idx < 57 and idx > 25:	
				wb_total_s[col_a + col_b +str(cell_idx)] = ws['B' + str(idx)].value or "XXX"	
				print "A PRE: " + col_a + col_b + str(cell_idx) + " = " + "B" + str(idx)
				idx += 1

def scale_a_post():
	idx = 3		
	for col in char_range('D', 'Z'):
		if idx < 26 :
			wb_total_s['B' + col +str(cell_idx)] = ws['C' + str(idx)].value or "XXX"	
			print "AAA POST: " + 'B' + col + str(cell_idx) + " = " + "C" + str(idx)
			idx += 1

	for col_a in char_range('C', 'D'):
		for col_b in char_range('A', 'Z'):
			if idx < 57 and idx > 25:
				wb_total_s[col_a + col_b +str(cell_idx)] = ws['C' + str(idx)].value or "XXX"	
				print "AAA POST: " + col_a + col_b + str(cell_idx) + " = " + "C" + str(idx)
				idx += 1
	
	
def scale_b_pre():
	idx = 3		
	for col in char_range('F', 'Z'):
		if idx < 26 :	
			wb_total_s['D' + col+str(cell_idx)] = ws['E' + str(idx)].value or "XXX"	
			print "BBB PRE: " + 'D' + col + str(cell_idx) + " = " + "E" + str(idx)
			idx += 1

	for col in char_range('A', 'S'):
		wb_total_s['E' + col +str(cell_idx)] = ws['E' + str(idx)].value or "XXX"	
		print "BBB PRE: " + 'E' + col + str(cell_idx) + " = " + "E" + str(idx)
		idx += 1
	

def scale_b_post():
	idx = 3	
	for col in char_range('T', 'Z'):
		if idx < 10 :
			wb_total_s['E' + col +str(cell_idx)] = ws['F' + str(idx)].value or "XXX"	
			print "BBB POST: " + 'E' + col + str(cell_idx) + " = " + "F" + str(idx)
			idx += 1

	for col_a in char_range('F', 'G'):
		for col_b in char_range('A', 'Z'):
			if idx > 9 and idx < 43:
				wb_total_s[col_a + col_b +str(cell_idx)] = ws['F' + str(idx)].value or "XXX"	
				print "BBB POST: " + col_a + col_b + str(cell_idx) + " = " + "F" + str(idx)
				idx += 1

def scale_c_pre():
	idx = 3	
	for col in char_range('H', 'V'):
		if idx < 18 :
			wb_total_s['G' + col +str(cell_idx)] = ws['H' + str(idx)].value or "XXX"	
			print "CCC PRE: " + 'G' + col + str(cell_idx) + " = " + "H" + str(idx)
			idx += 1


def scale_c_post():
	idx = 3	
	for col in char_range('W', 'Z'):
		if idx < 7 :
			wb_total_s['G' + col +str(cell_idx)] = ws['I' + str(idx)].value or "XXX"	
			print "CCC POST: " + 'G' + col + str(cell_idx) + " = " + "I" + str(idx)
			idx += 1

	for col in char_range('A', 'K'):
		if idx > 6 :
			wb_total_s['H' + col +str(cell_idx)] = ws['I' + str(idx)].value or "XXX"	
			print "CCC POST: " + 'H' + col + str(cell_idx) + " = " + "I" + str(idx)
			idx += 1

def scale_d_pre():
	idx = 3	
	for col in char_range('L', 'Z'):
		# if idx < 18 :
		wb_total_s['H' + col +str(cell_idx)] = ws['K' + str(idx)].value or "XXX"	
		print "DDD PRE: " + 'H' + col + str(cell_idx) + " = " + "K" + str(idx)
		idx += 1

	for col in char_range('A', 'Y'):
		if idx > 17 :
			wb_total_s['I' + col +str(cell_idx)] = ws['K' + str(idx)].value or "XXX"	
			print "DDD PRE: " + 'I' + col + str(cell_idx) + " = " + "K" + str(idx)
			idx += 1


def scale_d_post():
	idx = 3	
	
	wb_total_s['IZ' +str(cell_idx)] = ws['L' + str(idx)].value or "XXX"	
	print "DDD POST: " + 'IZ' + str(cell_idx) + " = " + "L" + str(idx)
	idx += 1

	for col_a in char_range('J', 'K'):
		for col_b in char_range('A', 'Z'):
			if idx > 3 and idx < 43:
				wb_total_s[col_a + col_b +str(cell_idx)] = ws['L' + str(idx)].value or "XXX"	
				print "DDD POST: " + col_a + col_b + str(cell_idx) + " = " + "L" + str(idx)
				idx += 1


for path, dirs, files in os.walk(pwd):
	for file in files:
		if os.path.splitext(file)[1].lower() == '.xlsx':
		# 	if file.endswith(".xlsx"):


			filename = os.path.join(path, file)
			if 'Tool' in filename :
				print 'TOOOOOL : ' + os.path.splitext(file)[0]
				continue

			if 'tool' in filename :
				print 'TOOOOOL : ' + os.path.splitext(file)[0]
				continue

			# considering korean conditions

			wb = load_workbook(filename, data_only=True)
			ws = wb.worksheets[0]

			wb_total_s['A'+str(cell_idx)] = os.path.splitext(file)[0]

			scale_a_pre()
			scale_a_post()

			scale_b_pre()
			scale_b_post()

			scale_c_pre()
			scale_c_post()

			scale_d_pre()
			scale_d_post()

			cell_idx += 1
				
			wb_total.save("data_only.xlsx")
			print "COUNT" + str(count)
			